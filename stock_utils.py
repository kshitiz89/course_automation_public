import openpyxl
import os

def append_ohlc_to_excel(file_path, symbol, ohlc_list):
    if not os.path.exists(file_path):
        print(f"⚠️ Excel file not found: {file_path}")
        return

    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Find next available row after existing content
        next_row = sheet.max_row + 2
        sheet.cell(row=next_row, column=1).value = "OHLC Data"
        sheet.cell(row=next_row + 1, column=1).value = "Stock"
        sheet.cell(row=next_row + 1, column=2).value = "Date"
        sheet.cell(row=next_row + 1, column=3).value = "High"
        sheet.cell(row=next_row + 1, column=4).value = "Low"
        sheet.cell(row=next_row + 1, column=5).value = "Close"

        row_ptr = next_row + 2
        for ohlc in ohlc_list:
            sheet.cell(row=row_ptr, column=1).value = symbol
            sheet.cell(row=row_ptr, column=2).value = ohlc["Date"]
            sheet.cell(row=row_ptr, column=3).value = ohlc["High"]
            sheet.cell(row=row_ptr, column=4).value = ohlc["Low"]
            sheet.cell(row=row_ptr, column=5).value = ohlc["Close"]
            row_ptr += 1

        wb.save(file_path)
        print(f"✅ OHLC added for {symbol} in: {file_path}")
    except Exception as e:
        print(f"❌ Error updating Excel for {symbol}: {e}")
